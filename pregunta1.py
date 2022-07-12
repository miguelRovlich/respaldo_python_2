import openpyxl as op
def crear_parametros():
    load = op.load_workbook('Productos.xlsx')
    ex = load.active
    class nodo:
        
        def __init__(self, k):
            self.name = k
            self.vecinos = []
        def insertaVecino(self,n):
            if n not in self.vecinos:
                self.vecinos.append(n)
    class grafo:
        def __init__(self):
            self.nodos = {}
        def insertaNodo(self,n):
            if n not in self.nodos:
                self.nodos[n] = nodo(n)
        def insertaArista(self,i,j):
            if i in self.nodos and j in self.nodos:
                self.nodos[i].insertaVecino(j)
                self.nodos[j].insertaVecino(i)
    #Listas donde se guardará cada requerimiento             
    lista_nodos=[]
    aristas=[]
    pesos=[]
    #Ingreso de requerimientos        
    no = int(input('Número de nodos: '))
    for i in range(no):
        ver=input('Ingrese nombre del nodo: ')
    #Agregar cada vertice ingresado           
        lista_nodos.append(ver)
    a = int(input('Número de aristas: '))
    #Se arma grafo a partir de nodo de inicio y destino para cada arista existente, una por una con su peso correspondiente
    for i in range(a):
        ni=input('Ingrese nombre del nodo de inicio: ')
        nd=input('Ingrese nombre del nodo de destino: ')
        dvert=input('Ingrese peso de cada arista: ')
    #Agregar cada arista y peso ingresado        
        aristas.append([ni,nd])
        pesos.append(dvert) 
    #Se crea el diccionario
    g = grafo()  
    #Se interactua este el diccionario creado y la función
    for i in lista_nodos:
        g.insertaNodo(i)
    for i in aristas:
        for e in i:
            g.insertaArista(i[0],i[1])
    for i in g.nodos:
        print(i,g.nodos[i].vecinos)
    #Agregar a excel los nodos, aristas y pesos         
    for i in range(len(lista_nodos)):
        ex.cell(1,i+2,lista_nodos[i])
        ex.cell(i+2,1,lista_nodos[i])
    for i in range(len(lista_nodos)):
        for e in range(len(lista_nodos)):
            ari= [lista_nodos[i],lista_nodos[e]]
            ari_r=[lista_nodos[e],lista_nodos[i]]
            if ari in aristas:
                ex.cell(i+2,e+2,int(pesos[aristas.index(ari)]))
            elif ari_r in aristas:
                ex.cell(i+2,e+2,int(pesos[aristas.index(ari_r)]))
            else:
                ex.cell(i+2,e+2,0)
    #Guardar excel con matriz                    
    load.save('Productos.xlsx')
    load.close()








opcion = 0
while opcion != 4:
    print("1. Crear parametros")
    print("2. resolver el algoritmo de Dijkstra")
    print("3. resolver el algoritmo de Kruskal")
    print("4. salir")
    opcion = int(input("ingrese una opcion: "))
    if opcion == 1:
        print("resolver el algoritmo de dijkstra")
    elif opcion == 2:
        print("resolver el algoritmo de floyd")
    elif opcion == 3:
        print("resolver el algoritmo de prim")
    elif opcion == 4:
        print("salir")
    else:
        print("opcion incorrecta, intente nuevamente")